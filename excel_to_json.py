from openpyxl import load_workbook
from collections import OrderedDict
import json

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
load_wb = load_workbook("/Users/dorong/Programming/Python/notion_api_wordbook/3000words.xlsx", data_only=True)
# 시트 이름으로 불러오기 
load_ws = load_wb['Sheet1']

c1 = '영단어'
c2 = '단어뜻'
c3 = '수준'

max_row = load_ws.max_row
max_col = load_ws.max_column
data_list = []
for row in range(1, max_row + 1):
    data = OrderedDict()
    for col in range(1, max_col + 1):
        if col == 1:
            data[c1] = load_ws.cell(row, col).value
        elif col == 2:
            data[c2] = load_ws.cell(row, col).value
        else:
            data[c3] = load_ws.cell(row, col).value
    data_list.append(data)

with open("/Users/dorong/Programming/Python/notion_api_wordbook/new_word.json", "w", encoding="utf-8") as f:
    json.dump(data_list, f, indent=4, ensure_ascii=False, default=str)